# app.py
import streamlit as st
from pathlib import Path
import tempfile
import datetime
from zoneinfo import ZoneInfo
import pandas as pd
import io

# Project utilities (assumes these files exist under src/)
from src.io_utils import read_products, read_schedule, allowed_file, check_size
from src.processor import generate_branch_date_files, create_zip_from_paths
from src.utils import ensure_category_column

st.set_page_config(page_title="Inventory / Stock Counting", layout="wide")
st.title("Inventory Application — Stock Counting Reports")
st.markdown(
    "Upload a **Products** file and a **Counting Schedule** file. "
    "The app will generate Excel reports for *today* (Africa/Cairo). "
    "Reports include blank `actual_quantity` cells and formula-driven `difference` cells so your team can fill actuals in Excel and see live results."
)

# File uploaders
uploaded_products = st.file_uploader(
    "Upload Products File (CSV, XLS, XLSX) — max 200MB",
    type=['csv', 'xls', 'xlsx'],
    key="products_uploader"
)
uploaded_schedule = st.file_uploader(
    "Upload Schedule File (CSV, XLS, XLSX) — max 200MB",
    type=['csv', 'xls', 'xlsx'],
    key="schedule_uploader"
)

# Helper: current date in Africa/Cairo
def get_today_cairo():
    try:
        tz = ZoneInfo("Africa/Cairo")
    except Exception:
        tz = datetime.timezone.utc
    return datetime.datetime.now(tz).date()

# -------------------------
# PRODUCTS preview & diagnostics
# -------------------------
st.header("Products preview / diagnostics")
if uploaded_products:
    st.write("Filename:", uploaded_products.name)
    ok, msg = allowed_file(uploaded_products)
    if not ok:
        st.error(msg)
    else:
        size_ok, size_msg = check_size(uploaded_products, max_mb=200)
        if not size_ok:
            st.error(size_msg)
        else:
            # Attempt to use read_products preview API (which may return a DataFrame)
            try:
                try:
                    uploaded_products.seek(0)
                except Exception:
                    pass
                preview_result = read_products(uploaded_products, preview=True)
                # If read_products returns DataFrame -> show it
                if isinstance(preview_result, pd.DataFrame):
                    if preview_result.empty:
                        st.info("Products preview returned an empty DataFrame (no rows found).")
                    else:
                        st.dataframe(preview_result)
                        st.success("Products file loaded for preview.")
                else:
                    # preview_result might be an iterator or other; try getting a first chunk
                    try:
                        uploaded_products.seek(0)
                    except Exception:
                        pass
                    it = read_products(uploaded_products, preview=False)
                    try:
                        first_chunk = next(it)
                        if isinstance(first_chunk, pd.DataFrame) and not first_chunk.empty:
                            st.dataframe(first_chunk.head(10))
                            st.success("Products file loaded (first chunk preview).")
                        else:
                            st.info("Products reader returned a chunk but it was empty.")
                    except StopIteration:
                        st.info("Products iterator yielded no chunks (empty file).")
            except Exception as e:
                st.warning(f"Automatic products preview failed: {e}")
                # Manual diagnostics: show raw head or sheet names
                try:
                    uploaded_products.seek(0)
                except Exception:
                    pass
                name_low = uploaded_products.name.lower()
                if name_low.endswith(".csv"):
                    try:
                        uploaded_products.seek(0)
                        raw = uploaded_products.read(5000)
                        try:
                            text = raw.decode(errors='replace')
                        except Exception:
                            text = str(raw)
                        st.code(text[:4000], language='text')
                        uploaded_products.seek(0)
                        try:
                            df_try = pd.read_csv(uploaded_products, nrows=10, dtype=str)
                            st.write("Detected columns (csv):", list(df_try.columns))
                            st.dataframe(df_try.head(10))
                        except Exception:
                            uploaded_products.seek(0)
                            try:
                                df_try = pd.read_csv(uploaded_products, sep=';', nrows=10, dtype=str)
                                st.write("Detected columns with sep=';':", list(df_try.columns))
                                st.dataframe(df_try.head(10))
                            except Exception as e2:
                                st.error(f"CSV parsing failed: {e2}")
                    except Exception as e3:
                        st.error(f"CSV diagnostics failed: {e3}")
                else:
                    # Excel diagnostics
                    try:
                        from openpyxl import load_workbook
                        uploaded_products.seek(0)
                        wb = load_workbook(filename=uploaded_products, read_only=True)
                        st.write("Excel sheets:", wb.sheetnames)
                    except Exception:
                        st.write("Could not list Excel sheets via openpyxl (file may be large or corrupted).")
                    try:
                        uploaded_products.seek(0)
                        xl = pd.ExcelFile(uploaded_products, engine='openpyxl')
                        for s in xl.sheet_names[:4]:
                            try:
                                df_try = xl.parse(s, nrows=5, engine='openpyxl', dtype=str)
                                st.write(f"Sheet '{s}' columns:", list(df_try.columns))
                                st.dataframe(df_try.head(5))
                            except Exception as e_sheet:
                                st.write(f"Could not parse sheet '{s}': {e_sheet}")
                    except Exception as e_xl:
                        st.error(f"Excel diagnostics failed: {e_xl}")
                st.error(
                    "Products preview could not be produced. Ensure the file has a header row with required columns "
                    "(name_en, branch_name, barcodes, brand, available_quantity) or upload as XLSX with the data sheet named 'data'."
                )
else:
    st.info("Upload a Products file to see a preview.")

# -------------------------
# SCHEDULE preview & diagnostics
# -------------------------
st.header("Schedule preview / diagnostics")
if uploaded_schedule:
    st.write("Filename:", uploaded_schedule.name)
    ok, msg = allowed_file(uploaded_schedule)
    if not ok:
        st.error(msg)
    else:
        size_ok, size_msg = check_size(uploaded_schedule, max_mb=200)
        if not size_ok:
            st.error(size_msg)
        else:
            # Try the read_schedule preview (it returns small df or raises helpful error)
            try:
                try:
                    uploaded_schedule.seek(0)
                except Exception:
                    pass
                sched_preview = read_schedule(uploaded_schedule, preview=True)
                if isinstance(sched_preview, pd.DataFrame) and not sched_preview.empty:
                    st.dataframe(sched_preview)
                    st.success("Schedule loaded for preview.")
                else:
                    st.info("Schedule preview returned empty result (no parsable rows found).")
            except Exception as e:
                st.warning(f"Automatic schedule preview failed: {e}")
                try:
                    uploaded_schedule.seek(0)
                except Exception:
                    pass
                name_low = uploaded_schedule.name.lower()
                if name_low.endswith(".csv"):
                    try:
                        uploaded_schedule.seek(0)
                        raw = uploaded_schedule.read(4000)
                        try:
                            text = raw.decode(errors='replace')
                        except Exception:
                            text = str(raw)
                        st.code(text[:3000], language='text')
                        uploaded_schedule.seek(0)
                        try:
                            df_try = pd.read_csv(uploaded_schedule, nrows=10, dtype=str)
                            st.write("Detected columns (csv):", list(df_try.columns))
                            st.dataframe(df_try.head(10))
                        except Exception:
                            uploaded_schedule.seek(0)
                            try:
                                df_try = pd.read_csv(uploaded_schedule, sep=';', nrows=10, dtype=str)
                                st.write("Detected columns with sep=';':", list(df_try.columns))
                                st.dataframe(df_try.head(10))
                            except Exception as e2:
                                st.error(f"CSV parsing failed: {e2}")
                    except Exception as e3:
                        st.error(f"CSV diagnostics failed: {e3}")
                else:
                    try:
                        from openpyxl import load_workbook
                        uploaded_schedule.seek(0)
                        wb = load_workbook(filename=uploaded_schedule, read_only=True)
                        st.write("Excel sheets:", wb.sheetnames)
                    except Exception:
                        st.write("Could not list Excel sheets via openpyxl.")
                    try:
                        uploaded_schedule.seek(0)
                        xl = pd.ExcelFile(uploaded_schedule, engine='openpyxl')
                        for s in xl.sheet_names[:4]:
                            try:
                                df_try = xl.parse(s, nrows=5, engine='openpyxl', dtype=str)
                                st.write(f"Sheet '{s}' columns:", list(df_try.columns))
                                st.dataframe(df_try.head(5))
                            except Exception as e_sheet:
                                st.write(f"Could not parse sheet '{s}': {e_sheet}")
                    except Exception as e_xl:
                        st.error(f"Excel diagnostics failed: {e_xl}")
                st.error(
                    "Schedule preview could not be produced. Ensure the schedule has headers including "
                    "branch, date, and brand (or acceptable alternatives) and that the header row is the first row."
                )
else:
    st.info("Upload a Schedule file to see a preview.")

# -------------------------
# Generate today's reports button
# -------------------------
st.header("Generate reports")
if st.button("Generate Today's Reports"):
    if not uploaded_products or not uploaded_schedule:
        st.error("Please upload BOTH Products and Schedule files before generating reports.")
    else:
        with st.spinner("Generating today's reports..."):
            try:
                # Read and expand schedule for today
                try:
                    uploaded_schedule.seek(0)
                except Exception:
                    pass
                schedule_df = read_schedule(uploaded_schedule, preview=False)  # full
                # Filter schedule to today (Africa/Cairo)
                today = get_today_cairo()
                schedule_today = schedule_df[schedule_df["date"] == today]
                if schedule_today.empty:
                    st.warning(f"No scheduled entries match today's date ({today}). No reports generated.")
                else:
                    st.write(f"Schedule rows matching today: {len(schedule_today)}")
                    st.dataframe(schedule_today.head(50))

                    # Read products (may be iterator)
                    try:
                        uploaded_products.seek(0)
                    except Exception:
                        pass
                    prod_preview = read_products(uploaded_products, preview=False)
                    # prod_preview is an iterator (yielding DataFrames). Use it directly in generator.
                    products_iter = prod_preview

                    # Create temp output dir
                    out_dir = Path(tempfile.mkdtemp(prefix="stock_reports_"))
                    generated_files = generate_branch_date_files(products_iter, schedule_today, out_dir)

                    if not generated_files:
                        st.warning("No reports were generated. The app did not find matching products for the scheduled branches/brands.")
                    else:
                        # create zip
                        zip_path = out_dir / f"Stock_Reports_{today.strftime('%d-%m-%Y')}.zip"
                        create_zip_from_paths(generated_files, zip_path)
                        st.success(f"Generated {len(generated_files)} file(s).")
                        with open(zip_path, "rb") as f:
                            st.download_button(
                                label="Download Today's Reports (ZIP)",
                                data=f,
                                file_name=zip_path.name
                            )
            except Exception as e:
                st.error(f"An error occurred during processing: {e}")

# -------------------------
# Footer / tips
# -------------------------
st.write("---")
st.caption(
    "Notes: • The app expects product headers like name_en, branch_name, barcodes, brand, available_quantity. "
    "If actual_quantity is missing it will be created blank in the generated files. • The schedule must contain branch, date, brand. "
    "Dates may be Excel serials or day numbers (1–31), the app uses Africa/Cairo timezone for 'today'."
)
